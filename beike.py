import requests
import urllib.request as r
from scrapy.selector import Selector
import re
import openpyxl
from openpyxl.styles import Font

def web(x,x1):
    url = 'https://cd.ke.com/ershoufang/{}/pg{}/'.format(x,x1)
    data=r.urlopen(url).read().decode('utf-8','ignore')
    return re.compile('href="(.*?)" target="_blank" title=').findall(data)[:30]

def fangwuxinxi(url):
    req=r.Request(url,headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36'})#伪装
    data=r.urlopen(req).read().decode('utf-8','ignore')
    selector=Selector(text=data)
    list2=[]
    try:
        s1=selector.xpath('//*[@id="beike"]/div[1]/div[2]/div[2]/div/div/div[1]/h1/@title').extract()[0]#售房名称
        s2=selector.xpath('//*[@id="beike"]/div[1]/div[4]/div/div[2]/div[4]/div[2]').xpath('string(.)').extract()[0].replace('\n','').replace(' ','')#所在区域
        s3=selector.xpath('//*[@id="beike"]/div[1]/div[4]/div/div[2]/div[4]/div[1]').xpath('string(.)').extract()[0].replace('\n','').replace(' ','')#所在小区
        s4=selector.xpath('//*[@id="introduction"]/div/div/div[1]/div[2]/ul/li[1]').xpath('string(.)').extract()[0]#房屋户型
        s5=selector.xpath('//*[@id="introduction"]/div/div/div[1]/div[2]/ul/li[3]').xpath('string(.)').extract()[0]#建筑面积
        s6=selector.xpath('//*[@id="beike"]/div[1]/div[4]/div/div[2]/div[2]/div[1]/div[1]').xpath('string(.)').extract()[0].replace('\n','').replace(' ','')#元/平方米
        s7=selector.xpath('//*[@id="beike"]/div[1]/div[4]/div/div[2]/div[2]/span[1]').xpath('string(.)').extract()[0]+'万'#房子总价
        s8=selector.xpath('//*[@id="zuanzhan"]/div[2]/div/div[1]/div[2]/div[1]/a').xpath('string(.)').extract()[0].replace(' ','')#房屋中介
        s9=selector.xpath('//*[@id="zuanzhan"]/div[2]/div/div[2]/div[2]').xpath('string(.)').extract()[0].replace(' ','')#中介电话
        list2.append(s1);list2.append(s2);list2.append(s3);list2.append(s4);list2.append(s5);list2.append(s6);list2.append(s7);list2.append(s8);list2.append(s9)
    except IndexError:
        pass
    if len(re.compile('户型结构</span>(.*?)</li>',re.S).findall(data))==0:
        s10='-'
    else:
        s10=re.compile('户型结构</span>(.*?)</li>',re.S).findall(data)[0]
    if len(re.compile('所在楼层</span>(.*?)</li>',re.S).findall(data))==0:
        s11='-'
    else:
        s11=re.compile('所在楼层</span>(.*?)</li>',re.S).findall(data)[0]
    if len(re.compile('建筑类型</span>(.*?)</li>',re.S).findall(data))==0:
        s12='-'
    else:
        s12=re.compile('建筑类型</span>(.*?)</li>',re.S).findall(data)[0]
    if len(re.compile('房屋朝向</span>(.*?)</li>',re.S).findall(data))==0:
        s13='-'
    else:
        s13=re.compile('房屋朝向</span>(.*?)</li>',re.S).findall(data)[0]
    if len(re.compile('建筑结构</span>(.*?)</li>',re.S).findall(data))==0:
        s13='-'
    else:
        s14=re.compile('建筑结构</span>(.*?)</li>',re.S).findall(data)[0]
    if len(re.compile('装修情况</span>(.*?)</li>',re.S).findall(data))==0:
        s15='-'
    else:
        s15=re.compile('装修情况</span>(.*?)</li>',re.S).findall(data)[0]
    if len(re.compile('梯户比例</span>(.*?)</li>',re.S).findall(data))==0:
        s16='-'
    else:
        s16=re.compile('梯户比例</span>(.*?)</li>',re.S).findall(data)[0]
    if len(re.compile('配备电梯</span>(.*?)</li>',re.S).findall(data))==0:
        s17='-'
    else:
        s17=re.compile('配备电梯</span>(.*?)</li>',re.S).findall(data)[0]
    list2.append(s10);list2.append(s11);list2.append(s12);list2.append(s13);list2.append(s14);list2.append(s15);list2.append(s16);list2.append(s17)
    return list2
print(web('longquanyi',1))
wb = openpyxl.Workbook()
sheet= wb['Sheet']
list1=["售房名称","所在区域","所在小区","房屋户型","建筑面积","元/平方米","房子总价","房屋中介","中介电话","户型结构","所在楼层","建筑类型","房屋朝向","建筑结构","装修情况","梯户比例","配备电梯"]
j=1
shuru=str(input('请输入爬取区域:(拼音)'))
shuru1=int(input('请输入爬取页数:'))
p=0
for i in range(1,18):
    cell=sheet.cell(row=1,column=i,value=list1[i-1])
for i in range(1,shuru1+1):#爬取区间
    for n in range(30):#爬取条数
        list3=fangwuxinxi(web(shuru,i)[n])
        j+=1
        for k in range(1,18):#写入售房信息
            p+=1
            cell=sheet.cell(row=j,column=k,value=list3[k-1])
            print('已爬取{}'.format(p))
        wb.save('b.xlsx')